from openpyxl import load_workbook
import sys
from collections import Counter

#import xlsx file
#xls_file = input('Enter Excel file path')
xls_file = sys.argv[1]
def cell_value(cell):
    return cell.value
workbook = load_workbook(xls_file)
worksheet = workbook['Cerrados']
print(f'Worksheet: {worksheet.title}')
print(f'Columnas: {worksheet.max_column}')
print(f'Filas: {worksheet.max_row}')
titles = next(worksheet.rows)
for title in titles:
    #print(f'{title.value}, {title.column_letter}{title.row}')
    if (title.value.lower() == 'estado'):
        print(f'Columna estado corresponde a letra {title.column_letter}')
        status_col = list(worksheet[title.column_letter])
        status_col.pop(0)
        status = map(cell_value, status_col)
        #print(f'Existen {len(status_col)} tickets {worksheet.title}, con estado:')
        #for col in status:
        #    print(col)
        counts = Counter(status)
        for value, count in counts.items():
            print(f'{value}: {count}')
    if (title.value.lower() == 'tipo'):
        print(f'Columna tipo corresponde a letra {title.column_letter}')
        type_col = list(worksheet[title.column_letter])
        type_col.pop(0)
        type = map(cell_value, type_col)
        counts = Counter(type)
        for value, count in counts.items():
            print(f'{value}: {count}')
    #    print(f'Existen {len(type_col)} tickets {worksheet.title} de tipo:')
    #    for col in type_col:
    #        print(col.value)

    




#for worksheet in worksheets:
#    workbook.active = worksheet
#    print(f'Worksheet: {workbook.active.title}')
#    columns = list(workbook.active.iter_cols(max_row=1))
#    for column in columns:
#        cell, *other_cells = column
#        print(f'Cell {cell.column_letter}{cell.col_idx} = {cell.value}')


#Identify worksheets
#wsheets = wbook.worksheets

#for wsheet in wsheets:
#    print 
#    print('max row: ' + whseet.max_row)

#print('Worksheets:')
#for name in wbook.sheetnames:
#    print(name)

