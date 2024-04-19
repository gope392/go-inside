from openpyxl import load_workbook
import sys
from mapping import *
from classes import Ticket

#import xlsx file
#xls_file = input('Enter Excel file path')
xls_file = sys.argv[1]
workbook = load_workbook(xls_file)

tickets = []

for worksheet in workbook.worksheets:
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        ticket = Ticket(Ticket_Id=row[Ticket_Id],
                        Tipo=row[Tipo],
                        Resumen=row[Resumen],
                        Prioridad=row[Prioridad],
                        Estado=row[Estado],
                        Fecha_de_creacion=row[Fecha_de_creacion],
                        Fecha_de_cambio=row[Fecha_de_cambio],
                        Fecha_de_resolucion=row[Fecha_de_resolucion],
                        Resolucion=row[Resolucion],
                        Solicitante=row[Solicitante],
                        Asignado=row[Asignado],
                        Descripcion=row[Descripcion],
                        Fecha_estimada=row[Fecha_estimada],
                        Worksheet=worksheet.title)
        tickets.append(ticket)

def attr_count(attr, obj_list):
    attr_list = [getattr(obj, attr) for obj in obj_list]
    attr_counts={}
    for value in attr_list:
        attr_counts[value] = attr_counts.get(value, 0) + 1
    print(f'---{attr}---')
    for value, count in attr_counts.items():
        print(f'{value} : {count}')

attr_count('Tipo', tickets)
attr_count('Estado', tickets)
attr_count('Worksheet', tickets)
